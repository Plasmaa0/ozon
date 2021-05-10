try:
    import os
    import gc #garbage collector
    import sys
    import json
    import time
    import random
    import openpyxl
    import traceback
    from pprint import pprint
    from bs4 import BeautifulSoup
    from selenium import webdriver
    from fake_useragent import UserAgent
    from concurrent.futures import ThreadPoolExecutor
except:
    print('Modules not found.')
    exit(0)

'''
TODO 1: wrap crawler, parser results in dataclasses
#from dataclasses import dataclass

TODO 2: make Scraper.__log() to be a decorator to catch and log exceptions
'''

class Scraper:
    def __init__(self, mode, max_threads = 20, save_xlsx = True, save_json = False, saveper = 30, maxattempts=3, loglevel=1, logcolor=False):
        '''
        Scraper class\n
        @max_threads: int - maximum number of threads used by scraper\n
        @save_xlsx: bool - save the xlsx table if True\n
        @save_json: bool - save the json if True\n
        @saveper: int - size of temporary file\n
        @maxattempts - maximum number of attempts to access the web-page\n
        @loglevel - loglevel:\n
            0 - NONSET\n
            1 - INFO\n
            2 - MESSAGE\n
            3 - WARNING\n
            4 - ERROR\n
            5 - CRITICAL\n
        @logcolor - if your terminal supports colored output.
        '''
        self.mode = mode
        self.__max_threads = max_threads
        self.__save_json = save_json
        self.__save_xlsx = save_xlsx
        self.__saveper = saveper
        self.__maxattempts = maxattempts
        self.__save_Failure = False
        self.__loglevel = loglevel
        self.__logcolor = logcolor
        self.__logfile = open('logs/log_{}.txt'.format(time.ctime().replace(' ', '_').replace('__','_').replace(':','-')), 'w')
        self.__uagent = UserAgent()
        self.__failurestatus = 'Failure'
        self.__successstatus = 'Success'
        self.dataset = []
        try:
            with open('stats.json', 'r') as f:
                self.__log('Accessed "stats.json"', 1)
                stats = json.load(f)
                try:
                    self.__logcolor = stats['color']
                except:
                    self.__log('no "color" field found in "stats.json"')
                    self.__logcolor = False
        except:
            self.__log('Not found "stats.json" to set log color. Recommended to run log color test.', 3)
            self.__logcolor = False
        else:
            self.__log('log color set to {}'.format(self.__logcolor), 1)
        self.__log('Scraper initialized successfully', 3)

    def __del__(self):
        self.__log('Closing Scraper', 3)
        self.__log('Closing processes', 3)
        self.__closeprocesses()
        self.__log('Running Garbage Collector', 3)
        gc.collect()
        self.__log('Detaching logfile', 3)
        self.__log('Finished', 3)
        self.__logfile.close()
        
    def __closeprocesses(self) -> None:
        '''
        Kills all 'chromedriver.exe' processes
        '''
        try:
            while True:
                r1 = os.system('TASKKILL /IM chromedriver.exe /f') 
                r2 = os.system('TASKKILL /IM chrome.exe /f')
                if r1!=0 and r2!=0:
                    break
        except Exception as e:
            self.__log('Error in __closeprocesses.' + str(e.args), 5)
        else:
            self.__log('Successfully closed processes', 1)

    def __getsession(self) -> webdriver.Chrome:
        '''
        Opens chrome session in headless mode with random user agent
        '''
        try:
            options = webdriver.ChromeOptions()
            options.add_argument('user-agent={}'.format(self.__uagent.random))
            options.add_argument('--log-level=3')
            options.add_argument('--disable-extensions')
            options.add_experimental_option('excludeSwitches', ['enable-logging'])
            options.add_argument('--no-sandbox')
            options.headless = True
            try:
                session = webdriver.Chrome(r'/usr/bin/chromedriver', options=options)
            except:
                session = webdriver.Chrome(r'/usr/local/bin/chromedriver', options=options)
            self.__log('Successfully got session', 0)
            return session
        except Exception as e:
            self.__log('Error in __getsession.' + str(e.args), 5)

    def __id2url(self, id_: int) -> str:
        '''
        Converting id of ozon seller to it's url\n
        id_ = 10213 --> 'https://www.ozon.ru/seller/10213/'
        '''
        try:
            url = 'https://www.ozon.ru/seller/{}/'.format(id_)
            self.__log('Converted url', 0)
            return url
        except Exception as e:
            self.__log('Error in __id2url.' + str(e.args), 5)
    
    def __pr(self, id, n):
        try:
            a = str(id)
            a = a + ' '*(n+4-len(a))
            self.__log('Formatted string for printing', 0)
            return a
        except Exception as e:
            self.__log('Error in __pr.' + str(e.args), 5)

    def __crawl(self, id_: int) -> dict:
        '''
        Crawls page with given id
        '''
        try:
            session = self.__getsession()
            def failuremessage(t1, t2, log, status):
                deltatime = t2-t1
                self.__log(str(['crawler', id_, log]), -1)
                self.__log('ID {} crawled in {} seconds. Status: {}'.format(self.__pr(id_, 8), self.__pr(round(deltatime,2), 6), status), 4)
                return {
                    'id': id_,
                    'url': self.__id2url(id_),
                    'log': {
                        'messages': log,
                        'time': round(deltatime, 2)
                    },
                    'status': status
                    }
            t1 = time.time()
            log = []
            url = self.__id2url(id_)
            status = None
            attempts = 0
            while attempts <= self.__maxattempts:
                log.append('SentRequest ({} attempt)'.format(attempts))
                attempts += 1
                try:
                    response = session.get(url)
                    response = session.page_source
                except Exception as e:
                    log.append('FailureGetRequest')
                    status = self.__failurestatus
                    return failuremessage(t1, time.time(), log, status)
                else:
                    log.append('GotResponse')
                    if 'robots' in response.lower():
                        log.append('BlockedBySiteError(ROBOTS)')
                        session.close()
                        session.quit() #возможно здесь начинают плодиться сессии
                        session = self.__getsession()
                        sleeptime = random.randint(1,7)
                        log.append('Sleeping {} seconds'.format(sleeptime))
                        time.sleep(sleeptime)
                    else:
                        log.append('CrawledSuccessfully')
                        status = self.__successstatus
                        break
            else:
                log.append('MaximumAttemptsExceeded({})'.format(self.__maxattempts))
                status = self.__failurestatus
            if 'robots' in response.lower():
                status = self.__failurestatus
                return failuremessage(t1, time.time(), log, status)
            session.close()
            session.quit()
            deltatime = time.time()-t1
            result = {
                'id': id_,
                'markup': response,
                # 'log': {
                #     'messages': log,
                #     'time': round(deltatime, 2)
                # },
                'status': status
            }
            self.__log('ID {} crawled in {} seconds. Status: {}'.format(self.__pr(id_, 8), self.__pr(round(deltatime,2), 6), status), 4 if status==self.__failurestatus else 1)
            gc.collect()
            return result
        except Exception as e:
            self.__log('Error in __crawl().' + str(e.args), 5)

    def __parse(self, inputs: dict) -> dict:
        '''
        Parsing the given inputs from __crawl()
        '''
        try:
            t1 = time.time()
            url = self.__id2url(inputs['id'])
            name = None
            shopId = inputs['id']
            subscribers = None
            log = []
            status = inputs['status']
            if status == self.__successstatus:
                try:
                    markup = inputs['markup']
                except:
                    log.append('MarkupNotFound')
                else:
                    log.append('FoundMarkup')
                try:
                    soup = BeautifulSoup(markup, features='html.parser')
                except:
                    log.append('FailedToCreateSoup')
                else:
                    log.append('CreatedSoup')
                try:
                    rawdata = soup.find('div', id='__ozon').find_next('script')
                except Exception as e:
                    log.append('MAIN_JSON_NotFound')
                    status = self.__failurestatus
                    if self.__save_Failure:
                        with open('brokenpages/id{}.html'.format(shopId), 'w', encoding='utf-8') as f:
                            f.write(markup)
                            f.close()
                else:
                    log.append('Found_MAIN_JSON')
                    srcraw = rawdata.contents[0].replace('window.__NUXT__=JSON.parse(\'', '').replace('\');', '').replace('\\', '').replace('\"{', '{').replace('}\"', '}').split(',"shared"')[0] + '}' + '}'
                    if 'seo' in srcraw:
                        srcraw = srcraw.split(',"seo"')[0] + '}' + '}'
                    try:
                        src = json.loads(srcraw)
                    except Exception as e:
                        log.append('MAIN_JSONLoadError')
                        status = self.__failurestatus
                    else:
                        log.append('Loaded_MAIN_JSON')
                        try:
                            div_id = 'state-' + src['state']['layout'][4]['placeholders'][0]['widgets'][0]['stateId']
                        except Exception as e:
                            log.append('MAIN_JSON_ParseError_1_attempt')
                            status = self.__failurestatus
                            try:
                                div_id = 'state-' + src['state']['layout'][6]['placeholders'][0]['widgets'][0]['placeholders'][0]['widgets'][1]['placeholders'][0]['widgets'][0]['stateId']
                            except Exception as e:
                                log.append('MAIN_JSON_ParseError_2_attempt')
                                try:
                                    title = soup.find('title')
                                except:
                                    log.append('TitleNotFound')
                                    status = self.__failurestatus
                                else:
                                    log.append('FoundTitle')
                                    try:
                                        name = (title.text.split('товары ')[1].split(' на')[0])
                                    except:
                                        log.append('NameNotFound')
                                        status = self.__failurestatus
                                    else:
                                        log.append('FoundName')
                                        status = self.__successstatus
                                finally:
                                    if 'ничего не нашлось' in markup.lower():
                                        log.append('Empty')
                                        status = self.__successstatus
                            else:
                                log.append('MAIN_JSON_ParseSuccess_2_attempt')
                                unparsed_data = soup.find('div', id=div_id).get('data-state')
                                try:
                                    unparsed_data = json.loads(unparsed_data)
                                except Exception as e:
                                    log.append('DATA_JSON_LoadError_2_attempt')
                                    status = self.__failurestatus
                                else:
                                    log.append('DATA_JSON_LoadSuccess_2_attempt')
                                    try:
                                        errmessage = unparsed_data['message']
                                    except:  
                                        log.append('DATA_JSON_2_attempt_ParseError') 
                                        try:
                                            title = soup.find('title')
                                        except:
                                            log.append('TitleNotFound')
                                            status = self.__failurestatus
                                        else:
                                            log.append('FoundTitle')
                                            try:
                                                name = (title.text.split('товары ')[1].split(' на')[0])
                                            except:
                                                log.append('NameNotFound')
                                            else:
                                                log.append('FoundName')
                                            status = self.__successstatus
                                    else:
                                        log.append('DATA_JSON_2_attempt_ParseSuccess')
                                        errmessage = errmessage.split('"')
                                        name = errmessage[1]
                                        log.append('Empty' if (('ничего не нашлось' in errmessage[2]) or ('ничего не нашлось' in markup.lower())) else '')
                                        status = self.__successstatus
                        else:
                            log.append('MAIN_JSON_ParseSuccess_1_attempt')
                            unparsed_data = soup.find('div', id=div_id).get('data-state')
                            try:
                                unparsed_data = json.loads(unparsed_data)
                            except Exception as e:
                                log.append('DATA_JSON_LoadError_1_attempt')
                                status = self.__failurestatus
                            else:
                                log.append('DATA_JSON_LoadSuccess_1_attempt')
                                try:
                                    name = unparsed_data['title']
                                except:
                                    log.append('NameNotFound')
                                else:
                                    log.append('FoundName')
                                try:
                                    subscribers = int(unparsed_data['subTitle'].split(' ')[0].replace('+', ''))
                                except:
                                    log.append('SubscribersNotFound')
                                else:
                                    log.append('FoundSubscribers')
                                status = self.__successstatus
                if name and subscribers:
                    log.append('Active')
                if name and 'ничего не нашлось' in (markup.lower()):
                    log.append('Empty')
                if name and not subscribers:
                    log.append('UnknownSubscribersCount')
                if 'Active' in log or 'Empty' in log or 'UnknownSubscribersCount' in log:
                    status = self.__successstatus
            else:
                log.append('IncorrectInput')
                status = self.__failurestatus
            deltatime = time.time()-t1
            data = {
                'name': name,
                'shopID': shopId,
                'url': url,
                'subscribers': subscribers,
                # 'log': {
                #     'crawler': inputs['log'],
                #     'parser': {
                #         'messages': log,
                #         'time': round(deltatime, 2)
                #     }
                # },
                'status': ('Active' if 'Active' in log else 'Empty') if status==self.__successstatus else 'Failure'
            }
            # self.__log(str(['crawler::::', sys.getsizeof(inputs)]), 5)
            # self.__log(str(['parser::::', sys.getsizeof(data)]), 5)
            if status == self.__failurestatus:
                self.__log(str(['parser', shopId, log]), -1)
            self.__log('ID {} parsed in {} seconds. Status: {}'.format(self.__pr(shopId, 8), self.__pr(round(deltatime,2), 6), status), 4 if status==self.__failurestatus else 1)
            gc.collect()
            return data
        except Exception as e:
            self.__log('Error in __parse().' + str(e.args), 5)

    def __handle(self, subindexes:list) -> None:
        '''
        Multithread Scraping all the id's in @subindexes list\n
        saving to self.dataset
        '''
        try:
            self.__log('Handling {} - {}.'.format(subindexes[0], subindexes[-1]), 2)
            with ThreadPoolExecutor(max_workers=self.__max_threads) as executor:
                self.__log('crawling '+ str(subindexes[0]) + '-' + str(subindexes[-1]), 2)
                markups = list(executor.map(self.__crawl, subindexes))
                self.__log('parsing ' + str(subindexes[0]) + '-' + str(subindexes[-1]), 2)
                data = list(executor.map(self.__parse, markups))
                self.dataset+=data
                executor.shutdown(True)
                self.__log('sleeping 5 sec', 1)
                time.sleep(5)
                self.__closeprocesses()
        except Exception as e:
            self.__log('Error in __handle().' + str(e.args), 5)
        else:
            self.__log('Handled succesfully {} - {}.'.format(subindexes[0], subindexes[-1]), 2)

    def __list2parts(self, list_:list, sizeOfPart:int):
        '''
        Splitting list to lists of the given size (@sizeOfPart)\n
        list(f([1,2,3,4,5,6,7], 2)) --> [[1,2],[3,4],[5,6],[7]]
        '''
        try:
            self.__log('list to parts', 0)
            for i in range(0, len(list_), sizeOfPart):
                yield list_[i:i + sizeOfPart]
        except Exception as e:
            self.__log('Error in __list2parts.' + str(e.args), 5)

    def __log(self, message, level):
        '''
        @level - loglevel:\n
            -1 - HIDDEN\n
            0 - NONSET (light-blue)\n
            1 - INFO (green)\n
            2 - MESSAGE (violet)\n
            3 - WARNING (yellow)\n
            4 - ERROR (orange)\n
            5 - CRITICAL (red)\n
        '''
        colors = {
            "NONSET": '\033[36m', #light-blue
            "INFO": '\033[32m', #green
            "MESSAGE": '\033[35m', #violet
            "WARNING": '\033[33m', #yellow
            "ERROR": '\033[34m', #orange
            "CRITICAL": '\033[31m' #red
        }
        if level == -1:
            log = 'HIDDEN'
        elif level == 0:
            log = "NONSET"
        elif level == 1:
            log = "INFO"
        elif level == 2:
            log = "MESSAGE"
        elif level == 3:
            log = "WARNING" 
        elif level == 4:
            log = "ERROR"
        elif level == 5:
            log = "CRITICAL"

        message = "<LOG. " + log + ". " + message
        if level != 0:
            print(time.ctime(), message+">", file=self.__logfile)
        if level>=self.__loglevel and level!=-1:
            if self.__logcolor:
                message = colors[log] + message + ">" + '\x1b[0m'
            else:
                message += ">"
            print(message)

    def timeConvert(self, n) -> str:
        '''
        Converting seconds to hours+minutes+seconds\n
        1245135.3 --> '345.0h : 52.0m : 15.30s'
        '''
        try:
            self.__log('converting time', 0)
            h=n//3600
            m=(n//60)%60
            s=n%60
            return "{}h : {}m : {:.2f}s".format(h,m,s)
        except Exception as e:
            self.__log('Error in timeConvert().' + str(e.args), 5)

    def scrape(self, a=1, b=1, definedIndexes = None):
        '''
        @a: int - Start index\n
        @b: int - End index\n
        @definedIndexes - if you want to scrape definded indxes range like [1,512,36455,445, ...]
        Starting process of scraping ozon sellers in range of start and end indexes\n
        Splitting all the range(a,b+1) to parts of size self.saveper for temporary save\n
        Calling __handle() to scrape all the parts and saving results\n
        Collecting all temporary results to a single file\n
        Deleting all temporary results
        '''
        try:
            assert (a and b) or definedIndexes
            self.__log('"xxx_xxx_tmp.json" files will be created during runtime\nDO NOT DELETE THEM!', 3)
            t0 = time.time()
            if definedIndexes:
                indexes = definedIndexes
                mode = self.mode
            else:
                indexes = [i for i in range(a,b+1)]
                mode = self.mode
            subindexes = list(self.__list2parts(indexes, self.__saveper))
            self.__log('working', 2)
            json_names = []
            xlsx_names = []
            parts = len(subindexes)
            for k, part in enumerate(subindexes):
                if 'stop.txt' in os.listdir():
                    self.__log('detected stop command. stopping scraping', 5)
                    os.remove('stop.txt')
                    break
                future = ThreadPoolExecutor(max_workers=1).submit(self.__handle, part)
                while future.running():
                    time.sleep(1)
                time.sleep(1)
                # self.__handle(part)
                subnames = self.save(True, self.__save_xlsx, part[0], part[-1], tmp=True)
                self.__log('scraped {} %'.format((k+1)*100/parts), 2)
                json_name, xlsx_name = subnames
                json_names.append(json_name)
                xlsx_names.append(xlsx_name)
                self.__log('saved from {} to {}'.format(part[0], part[-1]), 2)
                self.dataset.clear()
                gc.collect()
                time.sleep(2)
                
            self.__log('saving', 2)
            names = json_names + xlsx_names
            for name in json_names:
                with open(name, 'r', encoding='utf-8') as f:
                    self.dataset += json.load(f)
                    f.close()
            saved = self.save(self.__save_json, self.__save_xlsx, a, b, mode=mode)
            gc.collect()
            if saved:
                for name in names:
                    try:
                        os.remove(name)
                    except:
                        pass
            else:
                self.__log('Exception caught during save process. tmp files will not be deleted.', 4)
            self.__log('Total time: ' + self.timeConvert(round(time.time()-t0, 2)), 2)
            return saved
        except Exception as e:
            self.__log('Error in Scrape(). Raising it.' + str(e.args), 5)
            raise e

    def save(self, json_, xlsx_, a, b, mode='adding', tmp=False):
        '''
        Saving data in self.dataset in given extensions (json/xlsx)
        '''
        try:
            if mode == 'adding':
                finishTime = time.ctime()
                json_name = None
                xlsx_name = None
                if json_:
                    if not tmp:
                        self.__log('saving JSON file', 2)
                    json_name ='data/{}_{}data_{}.json'.format(a,b,finishTime.replace(' ', '_').replace(':','_'))
                    json_name = json_name if not tmp else 'data/{}_{}_tmp.json'.format(a,b)
                    with open(json_name, 'w', encoding='utf-8') as f:
                        json.dump(self.dataset, f, indent=4, ensure_ascii=False, sort_keys=True)
                        f.close()
                    if not tmp:
                        self.__log('saved JSON file "{}_{}data_{}.json"'.format(a,b,finishTime.replace(' ', '_').replace(':','_')), 2)
                if xlsx_ and not tmp:
                    if not tmp:
                        self.__log('saving XLSX file', 2)
                    wb = openpyxl.Workbook()
                    wb.create_sheet('shops', index=0)
                    sheet = wb['shops']
                    for row in range(len(self.dataset)):
                        shop = self.dataset[row]
                        keys = ['name', 'shopID', 'url', 'subscribers', 'status']
                        for k, key in enumerate(keys):
                            sheet.cell(row=1, column=k+1, value=key)
                        for col in range(len(keys)):
                            cell = sheet.cell(row=row+2, column=col+1, value=shop[keys[col]])
                            cell.value = shop[keys[col]]
                    xlsx_name = 'data/{}_{}data_{}.xlsx'.format(a,b,finishTime.replace(' ', '_').replace(':','_'))
                    xlsx_name = xlsx_name if not tmp else 'data/{}_{}_tmp.xlsx'.format(a,b)
                    wb.save(xlsx_name)
                    if not tmp:
                        self.__log('saved XLSL file "{}_{}data_{}.xlsx"'.format(a,b,finishTime.replace(' ', '_').replace(':','_')), 2)
                return [json_name, xlsx_name]
            elif mode == 'insertion':
                filename = None
                for file in os.listdir('data'):
                    if 'main' in file:
                        filename = file
                        break
                if filename == None:
                    self.__log('main table not found', 4)
                    wb = openpyxl.Workbook()
                    wb.create_sheet('shops', index=0)
                    sheet = wb['shops']
                    for row in range(len(self.dataset)):
                        shop = self.dataset[row]
                        keys = ['name', 'shopID', 'url', 'subscribers', 'status']
                        for k, key in enumerate(keys):
                            sheet.cell(row=1, column=k+1, value=key)
                    wb.save('data/1_1_data_main.xlsx')
                    filename = 'data/1_1_data_main.xlsx'
                    self.__log('created main table', 3)
                wb = openpyxl.load_workbook(filename if 'data/' in filename else 'data/'+filename)
                if 'data/' in filename:
                    filename = filename.replace('data/', '')
                sheet = wb['shops']
                id_ = 0
                updated = 0
                for shop in self.dataset:
                    keys = ['name', 'shopID', 'url', 'subscribers', 'status']
                    id_ = max(id_, shop['shopID'])
                    for col in range(len(keys)):
                        cell = sheet.cell(row=shop['shopID']+1, column=col+1, value=shop[keys[col]])
                        cell.value = shop[keys[col]]
                    updated += 1
                            # print(shop['shopID']+1, col+1, shop[keys[col]])
                        # print('\n')
                a,b = list(map(int, filename.split('_')[0:2]))[0:2]
                self.__log('uptaded {} rows in main table'.format(updated), 2)
                newfilename = '{}_{}_data_main.xlsx'.format(a,max(id_, b))
                try:
                    wb.save('data/'+newfilename)
                except:
                    self.__log('Exception caught during save process. Old table will not be deleted.', 4)
                else:
                    if newfilename != filename:
                        try:
                            os.remove('data/'+filename)
                        except:
                            pass
                return 1
        except Exception as e:
            self.__log('Error in save().' + str(e.args), 5)

    def test(self):
        try:
            n = int(input('Test sample size: '))
            t0 = time.time()
            files = self.scrape(1, 1, definedIndexes=sorted([random.randint(1,50000) for _ in range(n)]))
            total = time.time() - t0
            delta = total/100
            for file in files:
                try:
                    os.remove(file)
                except:
                    pass
            self.__log('Successfully ran test().', 3)
            return delta
        except Exception as e:
            self.__log('Error in test().' + str(e.args), 5)

class Program:
    def __init__(self):
        d = [i for i in os.walk('./')]
        if not 'data' in d[0][1]:
            print('folder "data" not found. creating it.')
            os.mkdir('data')
        if not 'logs' in d[0][1]:
            print('folder "logs" not found. creating it.')
            os.mkdir('logs')
        if not '3.4.3' in sys.version:
            print('Warning.\nYou\'re yousing python version:\n', sys.version, '\nand expected is 3.4.3')
        if not 'chromedriver.exe' in d[0][2]:
            print('no chromedriver found!!!!!!!!\n'*3)
            # raise FileNotFoundError
        if not ('stats.json' in d[0][2]):
            print('no "stats.json" found. recommended to run Test before scraping.')
        try:
            os.remove('stop.txt')
        except:
            pass

    def __catch(func):
        def wrapper(self):
            print('Starting tool: {}.'.format(func.__name__))
            try:
                assert callable(func)
                future = ThreadPoolExecutor(max_workers=1).submit(func, self)
                while future.running():
                    time.sleep(1)
                time.sleep(1)
                result = future.result()
                future.cancel()
                # result = func(self)
                return result
            except Exception as err:
                print('\nERROR!')
                try:
                    exc_info = sys.exc_info()
                finally:
                    traceback.print_exception(*exc_info)
                    print(err.args)
                    del exc_info
            finally:
                print('Closing tool: {}.'.format(func.__name__))
                print('{} finished running at {}'.format(func.__name__, time.ctime()))
                if 'currentsetup.json' in os.listdir():
                    print('deleting "currentsetup.json in 10 seconds"')
                    time.sleep(10)
                    try:
                        os.remove('currentsetup.json')
                    except:
                        pass
        return wrapper
    
    @__catch
    def comptmp(self):
        '''
        Compile all tmp files together
        '''
        print('Compiling tmp files!')
        names = sorted([i if 'tmp' in i else '' for i in os.listdir('data')])
        names = names[names.count(''):len(names)]
        if len(names) > 1:
            s = Scraper('adding', save_json=True)
            indexes = []
            for name in names:
                indexes += [int(i) for i in name.split('_')[:2]]
            for name in names:
                with open('data/'+name, 'r', encoding='utf-8') as f:
                    s.dataset += json.load(f)
                    f.close()
            s.save(input('Save json? (y/n) : ') == 'y', input('Save xlsx? (y/n) : ') == 'y', min(indexes), max(indexes))
            del s
            gc.collect()
            if input('Delete tmp files? (y/n): ') == 'y':
                for name in names:
                    os.remove('data/'+name)
    
    @__catch    
    def findFailures(self):
        print('Searching for failures!')
        files = os.listdir('data')
        limit = input('Limit of failures to find (Press Enter to find all failures): ')
        if limit.isnumeric():
            limit = int(limit)
        else:
            limit = None
        for file in files:
            if 'xlsx' in file:
                a,b = list(map(int, file.split('_')[0:2]))[0:2]
                wb = openpyxl.load_workbook('data/'+file, read_only=True)
                sheet = wb['shops']
                failed = []
                a = 0
                rows = [i for i in sheet.iter_rows()]
                for k, row in enumerate(rows):
                    if row[4].value == 'Failure':
                        a += 1
                        failed.append(row[1].value)
                    if a==limit:
                        break
                print('Found {} failures.'.format(len(failed)))
                if input('Want to scrape failed? (y/n): ') == 'y':
                    scr = Scraper('insertion')
                    n = len(failed)*self.getavgtime()
                    exp = scr.timeConvert(n)
                    if input('Expecting time: {}. Start? (y/n): '.format(exp)) == 'y':
                        scr.scrape(definedIndexes=failed)
                    elif input('Want to save failed in file? (y/n): ') == 'y':
                        print('saving failed indexes into file "data/{}_{}_faiures.json"'.format(a,b))
                        with open('data/{}_{}_faiures.json'.format(a,b), 'w', encoding='utf-8') as f:
                            json.dump(failed, f, indent=1, sort_keys=True)
                            f.close()
                    del scr
                    gc.collect()
                elif input('Want to save failed in file? (y/n): ') == 'y':
                    print('saving failed indexes into file "data/{}_{}_faiures.json"'.format(a,b))
                    with open('data/{}_{}_faiures.json'.format(a,b), 'w', encoding='utf-8') as f:
                        json.dump(failed, f, indent=1, sort_keys=True)
                        f.close()
                else:
                    pass

    @__catch
    def testcon(self):
        print('testing connection')
        scr = Scraper('adding', loglevel=2)
        delta = scr.test()
        del scr
        gc.collect()
        try:
            with open('stats.json', 'r') as f:
                d = json.load(f)
                f.close()
        except:
            print('Not found tests results to get avgtime. setting avgtime = 2')
            return 2
        else:
            try:
                colorsupport = d['color']
            except:
                colorsupport = None
                pass
            finally:
                with open('stats.json', 'w', encoding='utf-8') as f:
                    json.dump({
                        'AvgTime': delta+0.5,
                        'color': colorsupport
                    }, f,indent=4, ensure_ascii=False, sort_keys=True)
                    f.close()
                return delta

    @__catch
    def testcolor(self):
        try:
            with open('stats.json', 'r') as f:
                d = json.load(f)
                f.close()
        except:
            print('Not found tests results to check color support.')
        print('testing color support:')
        print('if the next string is "color" than your machine supports colored output')
        print("\033[31mc\x1b[0m" + "\033[32mo\x1b[0m" + "\033[33ml\x1b[0m" + "\033[34mo\x1b[0m" + "\033[35mr\x1b[0m")
        colorsupport = input('Is this string == "color"? (y/n): ') == 'y'
        try:
            delta = d['AvgTime']
        except:
            delta = None
        finally:
            with open('stats.json', 'w', encoding='utf-8') as f:
                        json.dump({
                            'AvgTime': delta,
                            'color': colorsupport
                        }, f,indent=4, ensure_ascii=False, sort_keys=True)
                        f.close()
            return colorsupport

    @__catch
    def clearlogs(self):
        count = 0
        for file in os.listdir('logs'):
            if 'log' in file:
                try:
                    os.remove('logs/'+file)
                except:
                    pass
                else:
                    count += 1
        print('Removed {} log files.'.format(count))

    @__catch
    def getavgtime(self):
        try:
            with open('stats.json', 'r') as f:
                d = json.load(f)
                f.close()
                assert d['AvgTime'] != None
                return d['AvgTime']
        except:
            print('Not found tests results to get avgtime. setting avgtime = 2')
            return 2

    @__catch
    def setup(self):
        '''
        Setup
        '''
        print('Scraper setup!')
        previousEnd = 0
        for f in os.listdir('data'):
            if 'main' in f:
                previousStart, previousEnd = list(map(int, f.split('_')[0:2]))[0:2]
                break
        if previousEnd != 0:
            print('Automatically found previous scraping End Index = {}. Setting it to be Start Index'.format(previousEnd))
            a = previousEnd + 1
        else:
            a = int(input('Start index: '))
            try:
                assert a > 0
            except:
                print('Start index must be bigger than 0')
                exit(0)
        if input('Select End index or delta from Start index? (1/2): ') == '1':
            b = int(input('End index: '))
            try:
                assert a <= b
            except:
                print('End index must be bigger or equal than Start index')
                exit(0)
        else:
            d = int(input('Delta from Start Index: '))
            b = a + d
        saveper = input('Save every ? results (Enter for recommended value): ')
        if saveper == '':
            print('Set recommended value')
            saveper = 30
        else:
            try:
                saveper = int(saveper)
                assert saveper > 10
            except:
                if saveper!= None:
                    if input('More than 10 results must be saved. Are you sure? (y/n): ')!='y':
                        exit(0)
                    elif type(saveper) != int:
                        raise TypeError('Wrong type of answer')
        maxattempts = input('Max request attempts (Enter for recommended value): ')
        if maxattempts == '':
            print('Set recommended value')
            maxattempts = 4
        else:
            try:
                maxattempts = int(maxattempts)
                assert maxattempts < 5
            except:
                if maxattempts != None:
                    if input('More than 5 attempts may make program too slow. Are you sure? (y/n): ')!='y':
                        exit(0)
                    elif type(maxattempts) != int:
                        raise TypeError('Wrong type of answer')
        max_thr = input('Max threads (Enter for recommended value): ')
        if max_thr == '':
            print('Set recommended value')
            max_thr = 20
        else:
            try:
                max_thr = int(max_thr)
                assert max_thr > 0
            except:
                if max_thr != None:
                    print('Max threads must be bigger than 0')
                    exit(0)
        if input('Insert results into main table or save into new files? (1/2): ') == '2':
            save_json = input('Save json? (y/n) : ') == 'y'
            save_xlsx = input('Save xlsx? (y/n) : ') == 'y'
            mode = 'adding'
            try:
                assert save_xlsx or save_json
            except:
                print('XLSX or JSON must be saved')
                exit(0)
        else:
            save_xlsx=False
            save_json=False
            mode = 'insertion'
        setup = {
            'setup': {
                'indexes': {
                    'start': a,
                    'end': b,
                    'delta': b-a
                },
                'max request attempts': maxattempts,
                'max threads': max_thr,
                'save': {
                    'JSON save': save_json,
                    'XLSX save': save_xlsx,
                    'mode': mode,
                    'tmp save size': saveper
                }
            }
        }
        scr = Scraper(mode, max_threads=max_thr, save_xlsx=save_xlsx, save_json=save_json, saveper=saveper, maxattempts=maxattempts)
        n = (1 + b - a)*self.getavgtime()
        exp = scr.timeConvert(n)
        pprint(setup, indent=2, compact=True, )
        if input('Expecting time: {}. Start? (y/n): '.format(exp)) == 'y':
            time.sleep(2)
            setup['started at'] = time.ctime()
            setup['expecting time'] = exp
            with open('currentsetup.json', 'w', encoding='utf-8') as f:
                print('Saving cuurent setup in "currentsetup.json".')
                json.dump(setup, f, indent=4)
                f.close()
            try:
                scr.scrape(a,b)
                del scr
                gc.collect()
            except Exception as err:
                print('ERROR!')
                try:
                    exc_info = sys.exc_info()
                finally:
                    traceback.print_exception(*exc_info)
                    print(err.args)
                    del exc_info
                    self.comptmp()
                    raise err
            finally:
                setup['finished at'] = time.ctime()
                with open('currentsetup.json', 'w', encoding='utf-8') as f:
                    json.dump(setup, f, indent=4)
                    f.close()

def main():
    prog = Program()
    print('Choose your tool !\n')
    choice = input('0.Exit\n1. Scraper\n2. Compile tmp files together\n3. Find failures\n4. Test connection (check what is average speed for parsing 1 page)\n5. Test color support\n6. Delete logs\n')
    if choice == '0':
        gc.collect()
        del prog
        return 1
    elif choice == '1':
        prog.setup()
    elif choice == '2':
        prog.comptmp()
    elif choice == '3':
        prog.findFailures()
    elif choice == '4':
        prog.testcon()
    elif choice == '5':
        prog.testcolor()
    elif choice == '6':
        prog.clearlogs()
    gc.collect()
    del prog
    return 0

if __name__ == '__main__':
    while True:
        future = ThreadPoolExecutor(max_workers=1).submit(main)
        while future.running():
            time.sleep(1)
        time.sleep(1)
        if future.result() == 1:
            break
        print('\n')
#
